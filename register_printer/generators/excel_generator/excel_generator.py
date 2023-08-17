import logging
import os.path
import sys
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from register_printer.constants import RW_TYPES, USER_VISIBLE_TYPES



LOGGER = logging.getLogger(__name__)

def get_excel_path(name):
    if hasattr(sys, 'frozen') and hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS,"register_printer/excels/" + name)
    return os.path.join(os.path.abspath("."),"register_printer/excels/" + name) 

def copy_cell_style(src_cell,tar_cell):
    tar_cell.value = src_cell.value
    if src_cell.has_style:
        tar_cell._style = copy(src_cell._style)
        tar_cell.font = copy(src_cell.font)
        tar_cell.border = copy(src_cell.border)
        tar_cell.fill= copy(src_cell.fill)
        tar_cell.number_format = copy(src_cell.number_format)
        tar_cell.protection = copy(src_cell.protection)
        tar_cell.alignment = copy(src_cell.alignment)

class ExcelGenerator:

    def __init__(self, top_sys, output_path="."):
        self.top_sys = top_sys
        self.output_path = output_path
        return

    def generate(self):
        LOGGER.debug("Generating Excel files...")

        self.generate_top()

        blocks_path = os.path.join(self.output_path, "blocks")

        os.makedirs(blocks_path, exist_ok=True)

        block_templates = self.top_sys.block_templates

        for block_template in block_templates:
            filename = os.path.join(
                blocks_path,
                '{0}_registers.xlsx'.format(
                    block_template.block_type
                )
            )
            ExcelGenerator.generate_block_template(
                filename,
                block_template
            )

        return

    @staticmethod
    def generate_block_template(
            filename,
            block_template):
        LOGGER.debug(
            "Generate file: %s for block template %s",
            filename,
            block_template.block_type)
        
        rb = load_workbook(get_excel_path('block.xlsx'))
        rs = rb.active

        wb = Workbook()
        ws = wb.active
        ws.page_setup = rs.page_setup
        ws.title = block_template.block_type

        for row in range(1,5):
            for col, value in enumerate(rs[row]):
                src_cell = rs.cell(row,col+1)
                tar_cell = ws.cell(row,col+1)
                copy_cell_style(src_cell,tar_cell)
        ws.merge_cells("A1:B1")
        ws.merge_cells("C1:H1")
        ws.merge_cells("A3:B3")
        ws.merge_cells("F4:H4")

        current_row = 1
        current_row += 4
        for array_template in block_template.array_templates:
            for col, vaule in enumerate(rs[4]):
                copy_cell_style(rs.cell(5,col+1),ws.cell(current_row,col+1))
            ws.merge_cells(f"F{current_row}:H{current_row}")
            ws.cell(current_row, 1).value = hex(array_template.offset)
            ws.cell(current_row, 2).value = array_template.name
            ws.cell(current_row, 3).value = array_template.length
            ws.cell(current_row, 4).value = hex(array_template.start_address)
            ws.cell(current_row, 5).value = hex(array_template.end_address)
            ws.cell(current_row, 6).value = str(array_template.description)
            current_row += 1

        current_row += 2
        copy_cell_style(rs.cell(7,1),ws.cell(current_row,1))
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws.auto_filter.ref = f"A{current_row+1}:M{current_row+1}"
        ws.freeze_panes = f"A{current_row+2}"

        current_row += 1
        for col, value in enumerate(rs[8]):
            copy_cell_style(rs.cell(8,col+1),ws.cell(current_row,col+1))
        current_row += 1

        valid_user_visible = DataValidation(
            type = 'list',
            formula1= f'\"{",".join(USER_VISIBLE_TYPES)}\"'
        )
        #valid_user_visible.promptTitle = 'user_visible'
        #valid_user_visible.prompt = 'please select Y,N, ,'
        #valid_user_visible.errorTitle = 'invalid user_visible'
        #valid_user_visible.error = 'Your user visible is not in the list'
        valid_access = DataValidation(
            type = 'list',
            formula1= f'\"{",".join(RW_TYPES)}\"'
        )
        ws.add_data_validation(valid_user_visible)
        ws.add_data_validation(valid_access)
        for register in block_template.register_templates:
            for col, vaule in enumerate(rs[8]):
                copy_cell_style(rs.cell(9,col+1),ws.cell(current_row,col+1))
                ws.cell(current_row,col+1).value = ''
            offset_cell = ws.cell(current_row, 1)
            offset_cell.value = hex(register.offset)
            name_cell = ws.cell(current_row, 2)
            name_cell.value = register.name
            description_cell = ws.cell(current_row, 8)
            description_cell.value = register.description
            current_row += 1
            hide_start = current_row
            for field in register.fields:
                for col, vaule in enumerate(rs[8]):
                    if field.user_visible == "Y":
                        if field.default != 0:
                            copy_cell_style(rs.cell(11,col+1),ws.cell(current_row,col+1))
                        else:
                            copy_cell_style(rs.cell(10,col+1),ws.cell(current_row,col+1))
                    else:
                        if field.default != 0:
                            copy_cell_style(rs.cell(13,col+1),ws.cell(current_row,col+1))
                        else:
                            copy_cell_style(rs.cell(12,col+1),ws.cell(current_row,col+1))
                    ws.cell(current_row,col+1).value = ''
                msb_cell = ws.cell(current_row, 3)
                msb_cell.value = field.msb
                lsb_cell = ws.cell(current_row, 4)
                lsb_cell.value = field.lsb
                field_name_cell = ws.cell(current_row, 5)
                field_name_cell.value = field.name
                access_cell = ws.cell(current_row, 6)
                valid_access.add(f"F{current_row}")
                access_cell.value = field.access
                default_value_cell = ws.cell(current_row, 7)
                default_value_cell.value = hex(field.default)
                field_description_cell = ws.cell(current_row, 8)
                field_description_cell.value = field.description
                user_visible_cell = ws.cell(current_row, 9)
                valid_user_visible.add(f"I{current_row}")
                user_visible_cell.value = field.user_visible
                description_chinese_cell = ws.cell(current_row, 10)
                description_chinese_cell.value = field.description_chinese
                current_row += 1
            # Add an empty line
            for col, vaule in enumerate(rs[8]):
                copy_cell_style(rs.cell(14,col+1),ws.cell(current_row,col+1))
            current_row += 1
            ws.sheet_properties.outlinePr.summaryBelow = False
            ws.row_dimensions.group(hide_start,current_row-2)
            hide_start = current_row + 1
        wb.save(filename=filename)

        return

    def generate_top(self):
        filename = os.path.join(
            self.output_path,
            "top.xlsx"
        )

        LOGGER.debug(
            "Generating Top excel to %s",
            filename)

        rb = load_workbook(get_excel_path('top.xlsx'))
        rs = rb.active

        wb = Workbook()
        ws = wb.active
        ws.title = "Top"

        ws.page_setup = rs.page_setup

        for row in range(1,8):
            for col, value in enumerate(rs[row]):
                src_cell = rs.cell(row,col+1)
                tar_cell = ws.cell(row,col+1)
                copy_cell_style(src_cell,tar_cell)

        ws["B1"] = self.top_sys.name
        ws["B2"] = self.top_sys.addr_width
        ws["B3"] = self.top_sys.data_width
        ws["B4"] = self.top_sys.author
        ws["B5"] = self.top_sys.version
        ws.auto_filter.ref = f"A7:F7"
        ws.freeze_panes = f"A8"

        # starting from row 8
        row = 8
        for block_instance in self.top_sys.block_instances:
            for col, value in enumerate(rs[8]):
                copy_cell_style(rs.cell(8+row%2,col+1),ws.cell(row,col+1))
            name_cell = ws.cell(row, 1)
            name_cell.value = block_instance.name
            type_cell = ws.cell(row, 2)
            type_cell.value = block_instance.block.block_type
            base_address_cell = ws.cell(row, 3)
            base_address_cell.value = hex(block_instance.base_address)
            size_cell = ws.cell(row, 4)
            size_cell.value = hex(block_instance.size)
            addr_width_cell = ws.cell(row, 5)
            raw_addr_width = block_instance.block.raw_addr_width
            if raw_addr_width is None:
                addr_width_cell.value = ""
            else:
                addr_width_cell.value = raw_addr_width
            data_width_cell = ws.cell(row, 6)
            raw_data_width = block_instance.block.raw_data_width
            if raw_data_width is None:
                data_width_cell.value = ""
            else:
                data_width_cell.value = raw_data_width
            row += 1
        wb.save(filename=filename)
        return
